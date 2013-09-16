from __future__ import division
#!/usr/bin/env python
# -*- coding: utf-8 -*-

#Third Party modules
import wx
import wx.html2
import openpyxl
import xlrd
import xlwt

#Python modules
import csv
import os
import time
import threading
import string
import shutil
import copy

#My Modules
import LinkBuilder
import pygmapscustom 
import Welcome

#Sets the location of the base map file
htmldir = os.getcwd()
htmlloc = htmldir + "\BaseMap.html"


class NonBlockingWorker(threading.Thread):
	def __init__(self, parent, command):
		threading.Thread.__init__(self)
		self.parent = parent
		self.command = command

	def run(self):
		if self.command == 'calculate':
			self.parent.Calculate(copy.copy(self.parent.ServiceLoc),self.parent.DemandLoc)
		elif self.command == 'Radius':
			self.parent.DistanceRadius()
			
		self.parent.CalcDistance.Enable()	
		self.parent.DistanceRadiusCalc.Enable()

#The main window in the gui
class MainFrame(wx.Frame):
	def __init__(self,parent,id,title):
		wx.Frame.__init__(self,parent,id,title,wx.DefaultPosition,style = wx.DEFAULT_FRAME_STYLE)
		
		#Declares initial variables
		self.ServicePath = ""
		self.ServiceLoc = {}
		self.DemandPath =  ""
		self.DemandLoc = {}
		self.FileAppendDict = {}
		self.TravelMethods = ["Driving","Bicycling","Walking"]
		self.TimeUnitList = ["Minutes","Hours"]
		self.DistUnitList = ["Kilometres","Miles"]
		self.fileselect = 0
		self.ChoicePostCol = 0
		self.list = []
		self.hexdict = {0:'00CC33',1:'19B72D',2:'33A328',3:'4C8E23',4:'667A1E',5:'667A1E',6:'995114',7:'B23D0F',8:'CC280A',9:'E51405',10:'FF0000'}
		self.MappingData = []
		self.DistRadDict = {}
		self.HeatDict = {}
		
		#The panel makes everything look similar on all platforms
		MainPanel =wx.ScrolledWindow(self,wx.ID_ANY)
		MainPanel.SetScrollbars(1, 1, 1, 1)
		MainPanel.EnableScrolling(True,True)
		
		self.browser = wx.html2.WebView.New(MainPanel,size= (1220,980),pos = (200,200))
		self.browser.LoadURL(htmlloc)
		ScrollWin = wx.ScrolledWindow(MainPanel,-1)
		#creates the GUI objects
		self.ServiceOpen = wx.Button(MainPanel,label = "Select service file",size = (190,114))
		self.DemandOpen = wx.Button(MainPanel,label = "Select demand file",size = (200,50))
		self.TravelMethod = wx.ComboBox(MainPanel,choices = self.TravelMethods, style = wx.CB_READONLY ,size = (200,22))
		
		self.UserSaveLoc = wx.CheckBox(MainPanel, label = "Choose output save location")
		self.CalcDistance = wx.Button(MainPanel,label = "Run",size = (420,50))
		self.DistGauge = wx.Gauge(MainPanel,range= 0,size = (320,50))
		self.Divider1 = wx.StaticLine(MainPanel,style=wx.LI_HORIZONTAL,size = (400,2))
		self.Divider2 = wx.StaticLine(MainPanel,style=wx.LI_HORIZONTAL,size = (400,2))
		self.Divider3 = wx.StaticLine(MainPanel,style=wx.LI_HORIZONTAL,size = (400,2))
		self.Divider4 = wx.StaticLine(MainPanel,style=wx.LI_HORIZONTAL,size = (400,2))
		self.DistanceRadiusCalc = wx.Button(MainPanel,label = "Run",size = (420,50))
		self.PostcodeColumnSelector = wx.ComboBox(MainPanel,choices = ["No file selected"],style = wx.CB_READONLY,size = (200,22))
		self.ChoosePostButt = wx.Button(MainPanel,label = "Load demand file",size = (420,50))
		self.ServiceRadInput = wx.TextCtrl(MainPanel,size = (75,-1))
		self.ServiceRadText = wx.StaticText(MainPanel,label = "Select demand nodes within ")
		self.DrivingRadInput = wx.TextCtrl(MainPanel,size = (50,-1))
		self.DrivingRadText = wx.StaticText(MainPanel,label = " of ")
		self.TimeUnitSelect = wx.ComboBox(MainPanel,choices = self.TimeUnitList ,style = wx.CB_READONLY,size = (200,22))
		self.DistUnitSelect = wx.ComboBox(MainPanel,choices = self.DistUnitList ,style = wx.CB_READONLY,size = (200,22))
		self.AppendOutputSelect = wx.CheckBox(MainPanel,label = "Append data to file")
		self.DistanceRadSlider = wx.Slider(MainPanel,-1,10,0,120,size = (400,-1),style = wx.SL_AUTOTICKS|wx.SL_LABELS)
		self.DistanceRadSlider.Disable()
		self.PostCodeColText = wx.StaticText(MainPanel,label = 'Postcode column')
		self.GaugeText = wx.StaticText(MainPanel,label = 'Status: Ready')
		self.MethodText = wx.StaticText(MainPanel,label = 'Mode of travel')
		self.TimeUnitText = wx.StaticText(MainPanel,label = 'Time units')
		self.DistUnitText = wx.StaticText(MainPanel,label = 'Distance units')
		self.MapTypeText = wx.StaticText(MainPanel,label = 'Map type')
		self.DemandSummary = wx.CheckBox(MainPanel,label = "Demand summary")
		self.DemandSummary.SetValue(True)
		self.MapOutput = wx.CheckBox(MainPanel,label = 'Map output')
		self.MapOutput.SetValue(True)
		self.RadUnitSelect = wx.ComboBox(MainPanel,choices = ['Minutes',"Kilometres","Miles"],style = wx.CB_READONLY)
		self.RadUnitSelect.SetValue('Minutes')
		self.CardiffLogo = wx.StaticBitmap(MainPanel)
		self.CardiffLogo.SetBitmap(wx.Bitmap("Logo.ico"))
		self.Bluei1 = wx.StaticBitmap(MainPanel)
		self.Bluei1.SetBitmap(wx.Bitmap("infoiconsmall.ico"))
		self.Bluei2 = wx.StaticBitmap(MainPanel)
		self.Bluei2.SetBitmap(wx.Bitmap("infoiconsmall.ico"))
		self.Bluei3 = wx.StaticBitmap(MainPanel)
		self.Bluei3.SetBitmap(wx.Bitmap("infoiconsmall.ico"))
		self.Bluei4 = wx.StaticBitmap(MainPanel)
		self.Bluei4.SetBitmap(wx.Bitmap("infoiconsmall.ico"))
		self.Bluei5 = wx.StaticBitmap(MainPanel)
		self.Bluei5.SetBitmap(wx.Bitmap("infoiconsmall.ico"))
		self.MapType = wx.ComboBox(MainPanel,choices = ["Standard","Heat Map"],style = wx.CB_READONLY,size = (200,22))
		self.MapType.SetValue("Standard")
		self.MapSave = wx.CheckBox(MainPanel,label = "Save map")
		
		self.ServiceEntryHeader1 = wx.StaticText(MainPanel, label = 'Postcode')
		self.ServiceEntryHeader2 = wx.StaticText(MainPanel, label = 'Name')
		self.ServiceEntryLoc1 = wx.TextCtrl(MainPanel)
		self.ServiceEntryName1 = wx.TextCtrl(MainPanel)
		self.ServiceEntryLabel1 = wx.StaticText(MainPanel, label = '1')
		self.ServiceEntryLoc2 = wx.TextCtrl(MainPanel)
		self.ServiceEntryName2 = wx.TextCtrl(MainPanel)
		self.ServiceEntryLabel2 = wx.StaticText(MainPanel, label = '2')
		self.ServiceEntryLoc3 = wx.TextCtrl(MainPanel)
		self.ServiceEntryName3 = wx.TextCtrl(MainPanel)
		self.ServiceEntryLabel3 = wx.StaticText(MainPanel, label = '3')
		self.ServiceEntryLoc4 = wx.TextCtrl(MainPanel)
		self.ServiceEntryName4 = wx.TextCtrl(MainPanel)
		self.ServiceEntryLabel4 = wx.StaticText(MainPanel, label = '4')
		self.ServiceEntryLoc5 = wx.TextCtrl(MainPanel)
		self.ServiceEntryName5 = wx.TextCtrl(MainPanel)
		self.ServiceEntryLabel5 = wx.StaticText(MainPanel, label = '5')
		
		HeaderFont = wx.Font(12,wx.DEFAULT,wx.NORMAL,wx.BOLD)
		self.DemandBlurb = wx.StaticText(MainPanel,label = 'Task 1: Demand')
		self.DemandBlurb.SetFont(HeaderFont)
		self.ServiceBlurb = wx.StaticText(MainPanel,label = 'Task 2: Service')
		self.ServiceBlurb.SetFont(HeaderFont)
		self.OptionsBlurb = wx.StaticText(MainPanel,label = 'Task 3: Options')
		self.OptionsBlurb.SetFont(HeaderFont)
		self.DistanceBlurb = wx.StaticText(MainPanel,label = 'Output 1 :Distance')
		self.DistanceBlurb.SetFont(HeaderFont)
		self.LimitBlurb = wx.StaticText(MainPanel,label = 'Output 2: Limit')
		self.LimitBlurb.SetFont(HeaderFont)
		
		#Sets the GUI objects to default
		self.TravelMethod.SetValue(self.TravelMethods[0])
		self.TimeUnitSelect.SetValue(self.TimeUnitList[0])
		self.DistUnitSelect.SetValue(self.DistUnitList[0])
		self.PostcodeColumnSelector.SetValue("No File Selected")
		
		#Binds events to objects
		self.Bind(wx.EVT_BUTTON,lambda event : self.Nonblocking('calculate'),self.CalcDistance)
		self.Bind(wx.EVT_BUTTON,self.IdService,self.ServiceOpen)
		self.Bind(wx.EVT_BUTTON,self.IdDemand,self.DemandOpen)
		self.Bind(wx.EVT_BUTTON,lambda event : self.Nonblocking('Radius'),self.DistanceRadiusCalc)
		self.Bind(wx.EVT_COMBOBOX,self.GetPostSel,self.PostcodeColumnSelector)
		self.Bind(wx.EVT_BUTTON,self.Rescan,self.ChoosePostButt)
		self.Bind(wx.EVT_SCROLL_CHANGED,self.ChangeBySlider, self.DistanceRadSlider)
		self.Bluei1.Bind(wx.EVT_LEFT_DOWN,lambda event : self.HelpWindow(1))
		self.Bluei2.Bind(wx.EVT_LEFT_DOWN,lambda event : self.HelpWindow(2))
		self.Bluei3.Bind(wx.EVT_LEFT_DOWN,lambda event : self.HelpWindow(3))
		self.Bluei4.Bind(wx.EVT_LEFT_DOWN,lambda event : self.HelpWindow(4))
		self.Bluei5.Bind(wx.EVT_LEFT_DOWN,lambda event : self.HelpWindow(5))		
		self.Bind(wx.EVT_CHECKBOX,self.MapTypeControl,self.MapOutput)
		
		#This is how the GUI is formatted
		vbox = wx.BoxSizer(wx.VERTICAL)
		hboxfunctions = wx.BoxSizer(wx.HORIZONTAL)
		vboxmap = wx.BoxSizer(wx.VERTICAL)
		vboxdistmatrix = wx.BoxSizer(wx.VERTICAL)
		hboxdm1 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdm2 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdm3 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdm4 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdm5 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdm6 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdm65 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdm675 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdm7 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdm8 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdr1 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdr2 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdr3 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdr4 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdr5 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdr6 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdr7 = wx.BoxSizer(wx.HORIZONTAL)
		hboxdr8 = wx.BoxSizer(wx.HORIZONTAL)
		hboxGaugeTXT = wx.BoxSizer(wx.HORIZONTAL)
		hboxGauge = wx.BoxSizer(wx.HORIZONTAL)
		hboxDiv1 = wx.BoxSizer(wx.HORIZONTAL)
		hboxDiv2 = wx.BoxSizer(wx.HORIZONTAL)
		hboxDiv3 = wx.BoxSizer(wx.HORIZONTAL)
		hboxDiv4 = wx.BoxSizer(wx.HORIZONTAL)
		vboxPCS = wx.BoxSizer(wx.VERTICAL)
		vboxGauge = wx.BoxSizer(wx.VERTICAL)
		hboxHeaderDemand = wx.BoxSizer(wx.HORIZONTAL)
		hboxHeaderService = wx.BoxSizer(wx.HORIZONTAL)
		hboxHeaderOptions = wx.BoxSizer(wx.HORIZONTAL)
		hboxHeaderDistance = wx.BoxSizer(wx.HORIZONTAL)
		hboxHeaderLimit = wx.BoxSizer(wx.HORIZONTAL)
		
		vboxSE = wx.BoxSizer(wx.VERTICAL)
		hboxSEH = wx.BoxSizer(wx.HORIZONTAL)
		hboxSE1 = wx.BoxSizer(wx.HORIZONTAL)
		hboxSE2 = wx.BoxSizer(wx.HORIZONTAL)
		hboxSE3 = wx.BoxSizer(wx.HORIZONTAL)
		hboxSE4 = wx.BoxSizer(wx.HORIZONTAL)
		hboxSE5 = wx.BoxSizer(wx.HORIZONTAL)
		
		vboxPCS.Add(self.PostCodeColText,flag = wx.ALIGN_CENTRE)
		vboxPCS.Add((0,5))
		vboxPCS.Add(self.PostcodeColumnSelector)
		
		hboxSEH.Add((20,0))
		hboxSEH.Add(self.ServiceEntryHeader1)
		hboxSEH.Add((53,0))
		hboxSEH.Add(self.ServiceEntryHeader2)
		hboxSE1.Add(self.ServiceEntryLabel1, wx.ALL,5)
		hboxSE1.Add((5,0))
		hboxSE1.Add(self.ServiceEntryLoc1)
		hboxSE1.Add(self.ServiceEntryName1)
		hboxSE2.Add(self.ServiceEntryLabel2, wx.ALL,5)
		hboxSE2.Add((5,0))
		hboxSE2.Add(self.ServiceEntryLoc2)
		hboxSE2.Add(self.ServiceEntryName2)
		hboxSE3.Add(self.ServiceEntryLabel3, wx.ALL,5)
		hboxSE3.Add((5,0))
		hboxSE3.Add(self.ServiceEntryLoc3)
		hboxSE3.Add(self.ServiceEntryName3)
		hboxSE4.Add(self.ServiceEntryLabel4, wx.ALL,5)
		hboxSE4.Add((5,0))
		hboxSE4.Add(self.ServiceEntryLoc4)
		hboxSE4.Add(self.ServiceEntryName4)
		hboxSE5.Add(self.ServiceEntryLabel5, wx.ALL,5)
		hboxSE5.Add((5,0))
		hboxSE5.Add(self.ServiceEntryLoc5)
		hboxSE5.Add(self.ServiceEntryName5)
		
		vboxSE.Add(hboxSE1)
		vboxSE.Add(hboxSE2)
		vboxSE.Add(hboxSE3)
		vboxSE.Add(hboxSE4)
		vboxSE.Add(hboxSE5)
				
		#Creates the Functions Box
		hboxdm2.Add(self.ChoosePostButt)
		
		hboxdm1.Add(self.DemandOpen)
		hboxdm1.Add((20,0))
		hboxdm1.Add(vboxPCS)
		
		hboxdm3.Add(hboxSEH)
		
		hboxdm4.Add(vboxSE)
		hboxdm4.Add((5,0))
		hboxdm4.Add(self.ServiceOpen)
		
		hboxdm5.Add(self.MethodText, wx.ALL,5)
		hboxdm5.Add((18,0))
		hboxdm5.Add(self.TravelMethod)
		
		hboxdm6.Add(self.TimeUnitText, wx.ALL,5)
		hboxdm6.Add((38,0))
		hboxdm6.Add(self.TimeUnitSelect)
		
		hboxdm65.Add(self.DistUnitText, wx.ALL,5)
		hboxdm65.Add((20,0))
		hboxdm65.Add(self.DistUnitSelect)
		
		hboxdm675.Add(self.MapTypeText,wx.ALL,5)
		hboxdm675.Add((42,0))
		hboxdm675.Add(self.MapType)
		
		hboxdm7.Add(self.CalcDistance)
		
		hboxdr1.Add(self.ServiceRadText,wx.TOP,5)
		hboxdr1.Add((5,0))
		hboxdr1.Add(self.DrivingRadInput)
		hboxdr1.Add((5,0))
		hboxdr1.Add(self.RadUnitSelect)
		hboxdr1.Add((5,0))
		vboxdummy = wx.BoxSizer(wx.VERTICAL)
		vboxdummy.Add((0,3))
		vboxdummy.Add(self.DrivingRadText)
		hboxdr1.Add(vboxdummy)
		hboxdr1.Add((5,0))
		hboxdr1.Add(self.ServiceRadInput)

		hboxdr3.Add(self.DistanceRadSlider)
		
		hboxdr4.Add(self.DistanceRadiusCalc)
		
		hboxdr2.Add(self.AppendOutputSelect)
		hboxdr2.Add((20,0))
		hboxdr2.Add(self.UserSaveLoc)
		
		vboxdum1 = wx.BoxSizer(wx.VERTICAL)
		vboxdum2 = wx.BoxSizer(wx.VERTICAL)
		vboxdum3 = wx.BoxSizer(wx.VERTICAL)
		
		vboxdum1.Add((0,5))
		vboxdum1.Add(self.DemandSummary,wx.TOP,5)
		hboxdr5.Add(vboxdum1)
		hboxdr5.Add((23,0))
		
		vboxdum2.Add(((0,5)))
		vboxdum2.Add(self.MapOutput,wx.TOP,5)
		hboxdr5.Add(vboxdum2)
		hboxdr5.Add((10,0))
		
		vboxdum3.Add((0,5))
		vboxdum3.Add(self.MapSave,wx.TOP,5)
		hboxdr5.Add(vboxdum3)
		
		hboxDiv1.Add(self.Divider1)
		hboxDiv2.Add(self.Divider2)
		hboxDiv3.Add(self.Divider3)
		hboxDiv4.Add(self.Divider4)
		
		vboxGauge.Add((0,5))
		vboxGauge.Add(self.GaugeText,flag = wx.ALIGN_LEFT)
		vboxGauge.Add((0,5))
		vboxGauge.Add(self.DistGauge)
		
		hboxdr6.Add(self.CardiffLogo,flag = wx.ALIGN_CENTRE)
		hboxdr6.Add((20,0))
		hboxdr6.Add(vboxGauge)
		
		hboxHeaderDemand.Add(self.DemandBlurb,flag = wx.ALIGN_CENTRE)
		hboxHeaderDemand.Add((10,0))
		hboxHeaderDemand.Add(self.Bluei1,wx.TOP,5)
		
		hboxHeaderService.Add(self.ServiceBlurb)
		hboxHeaderService.Add((10,0))
		hboxHeaderService.Add(self.Bluei2,wx.TOP,5)
		
		hboxHeaderOptions.Add(self.OptionsBlurb)
		hboxHeaderOptions.Add((10,0))
		hboxHeaderOptions.Add(self.Bluei3,wx.TOP,5)
		
		hboxHeaderDistance.Add(self.DistanceBlurb)
		hboxHeaderDistance.Add((10,0))
		hboxHeaderDistance.Add(self.Bluei4,wx.TOP,5)
		
		hboxHeaderLimit.Add(self.LimitBlurb)
		hboxHeaderLimit.Add((10,0))
		hboxHeaderLimit.Add(self.Bluei5,wx.TOP,5)
		
		#Formats the Functions Box
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxHeaderDemand,flag = wx.ALIGN_LEFT)
		vboxdistmatrix.Add((0,5))
		vboxdistmatrix.Add(hboxdm1, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,20))
		vboxdistmatrix.Add(hboxdm2, flag = wx.ALIGN_LEFT)
		vboxdistmatrix.Add((0,20))
		vboxdistmatrix.Add(hboxDiv1, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxHeaderService,flag = wx.ALIGN_LEFT)
		vboxdistmatrix.Add((0,5))
		vboxdistmatrix.Add(hboxdm3, flag = wx.ALIGN_LEFT)
		vboxdistmatrix.Add(hboxdm4, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,20))
		vboxdistmatrix.Add(hboxDiv2,flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxHeaderOptions,flag = wx.ALIGN_LEFT)
		vboxdistmatrix.Add((0,5))
		vboxdistmatrix.Add(hboxdr5,flag = wx.ALIGN_LEFT)
		vboxdistmatrix.Add((0,20))
		vboxdistmatrix.Add(hboxdr2,flag = wx.ALIGN_LEFT)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxDiv4)
		vboxdistmatrix.Add((0,5))
		vboxdistmatrix.Add(hboxHeaderDistance)
		vboxdistmatrix.Add((0,5))
		vboxdistmatrix.Add(hboxdm5, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxdm6, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxdm65, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxdm675, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,20))
		vboxdistmatrix.Add(hboxdm7, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,20))
		vboxdistmatrix.Add(hboxDiv3, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxHeaderLimit, flag = wx.ALIGN_LEFT)
		vboxdistmatrix.Add((0,5))
		vboxdistmatrix.Add(hboxdr1, flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxdr3,flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxdr4,flag = wx.ALIGN_CENTRE)
		vboxdistmatrix.Add((0,10))
		vboxdistmatrix.Add(hboxdr6,flag = wx.ALIGN_LEFT)

		vboxmap.Add(self.browser,flag = wx.EXPAND)
		
		#adds the objects to the window
		hboxfunctions.Add((20,0))
		hboxfunctions.Add(vboxdistmatrix)
		hboxfunctions.Add((20,0))
		hboxfunctions.Add(vboxmap, flag = wx.EXPAND)
		
		#vbox.Add(hboxfunctions)
		MainPanel.SetSizer(hboxfunctions)

	#All the functions for the main window
	def MapTypeControl(self,Event):
		if self.MapOutput.GetValue() == True:
			self.MapType.Enable()
		elif self.MapOutput.GetValue() == False:
			self.MapType.Disable()
	
	#Displays helptext on clicking info buttons
	def HelpWindow(self,InfoSelect):
		
		if InfoSelect == 1:
			Title = "Task 1 : Demand"
			Text = '''
			- Click on the Select demand file button to select the file where your demand locations are stored.
			This button will turn green once the file has been succesfully been selected.\n
			- Column headers will appear in the dropdown box under Postcode  column. 
			Select the column in which postcodes (or other address format) are stored.\n
			- Click on Load demand file. This button will turn green once the file has successfully been read.\n
			For more detailed information, see the PatMap User Guide
			'''

		elif InfoSelect == 2:
			Title = "Task 2: Service"			
			Text = '''
			- Enter postcodes and names of up to five service locations in the boxes provided.\n
			AND/OR\n
			-Click on the Select service file button to upload the file where your service locations are stored.
			The button will turn green once the file has successfully been read.\n
			For more detailed information, see the PatMap User Guide.
			'''
			
		elif InfoSelect == 	3:
			Title = "Task 3: Options"
			Text = '''
			- Tick required output options:\n
			- Demand Summary: Returns an Excel file with the number of times each postcode appears in your demand file.
			Time and distance between each demand and service location are also given.
			      - Map output: Updates map with demand and service locations.
			      - Save Map: Saves the map as a JPEG image.\n				  
			- Append data to file: Appends the time and distance to each service location to each observation listed in your demand file.
			This will automatically be saved in the same location as your demand file, withe _Append added to the demand file name.\n
			- Choose output save location: Choose the nameand location of the demand summary and map files. If this is not selected, by default they will be saved in the Testdata folder, located within the program folder,
			with a timestamp as the file names.\n
			For more detailed information, see the PatMap User Guide.
			'''
			
		elif InfoSelect == 	4:
			Title = "Output 1: Distance "
			Text = '''
			-Select mode of travel, time units and distance units required for your output using the dropdown
			boxes. (By default, these are driving, minuetes and kilometres respectively)\n
			-Select map type:\n
			Standard - each demand location is displayed as a regular icon on the map.\n
			Heat map - the colour assigned to each demand location icon is dependant upon demand density at that location.\n
			-Click on the run button.\n
			For more detailed information, see the PatMap User Guide.
			'''
			
		elif InfoSelect == 	5:
			Title = "Output 2: Limit"
			Text = '''
			- Enter the required numerical limit, select the units and enter the service location postcode.\n
			- Click on the Run button.
			- The map will display the demand locations according to whether or not they are within the specified limit. Those within the limit are coloured green,
			those outside the limit are coloured red.\n
			-Using the slider will dynamically update the map but not the output. To update the output, the options must be amended and the function run again.\n
			For more detailed information, see the PatMap User Guide.
			'''
			
		wx.MessageBox(Text, Title, wx.OK | wx.ICON_QUESTION)

	#Tells the program its choosing a demand file
	def IdService(self,event):
		self.fileselect = 1
		self.OpenFile()
		
	#Tells the program its choosing a demand file	
	def IdDemand(self,event):
		self.fileselect = 2
		self.OpenFile()
	
	#Prevents the GUI freezing when a function is called
	def Nonblocking(self,Command):
		#Disables the buttons the prevent multiple sending of the same data
		self.CalcDistance.Disable()
		self.DistanceRadiusCalc.Disable()
		worker = NonBlockingWorker(self,Command)
		worker.start()
		
	#Shows locations within a certain driving distance
	def DistanceRadius(self):
		#Initialises the variables
		self.DistanceRadSlider.Disable()
		ServiceRadLoc = self.ServiceRadInput.GetValue()
		ServiceRadLoc = ServiceRadLoc.strip()
		MappingData = []
		DemandWithin = []
		DemandOutside = []
		self.DistRadDict = {}
		self.DistRadAppend = {}
		
		#Ensures there is data to send
		if ServiceRadLoc == "" or len(ServiceRadLoc) > 12 or self.DemandLoc == {}:
			wx.MessageBox('Choose a service location and demand file', 'No postcodes!', wx.OK | wx.ICON_ERROR)
			return
			
		#Sets the Radius to default if nothing is entered
		if self.DrivingRadInput.GetValue() == "":
			DrivingRadius = 10
		else:
			try:
				DrivingRadius = eval(self.DrivingRadInput.GetValue())
			except:
				DrivingRadius = 10
		
		#Only runs the user selected options
		if self.DemandSummary.GetValue() == True or self.MapSave.GetValue() == True:
			
			#Chooses the save location
			if self.UserSaveLoc.GetValue() == False:
				filebase = ".\Testdata\Radius%s" %(time.time())
				filetype = ".csv"
				fileloc = filebase[:-3] + filetype
				
			else:
				fileloc = ""	
				SaveFileDialog = wx.FileDialog(self,"Choose Save Location", os.getcwd(),"" ,"csv Files (*.csv)|*.csv", wx.FD_SAVE)
			
				#Closes the dialog if the user cancels
				if SaveFileDialog.ShowModal() == wx.ID_CANCEL:
					return
				
				fileloc = SaveFileDialog.GetPath()
			Header = 'Within %s %s of %s' %(DrivingRadius,self.RadUnitSelect.GetValue(),ServiceRadLoc)
			if self.DemandSummary.GetValue() == True:
				#puts the header on the file
				outfile = open(fileloc,'ab')
				output=csv.writer(outfile)
				output.writerow(['Post Code',Header])
				outfile.close()
		
		#Stores the service Location
		Latlngdata = LinkBuilder.Geoencoding(ServiceRadLoc)
		
		if not Latlngdata[0] == "OK":
			wx.MessageBox('The service location has not been recognised by the google distance matrix service', 'Service postcode is invalid', wx.OK | wx.ICON_ERROR)
			return
		
		MappingData.append([str(ServiceRadLoc),Latlngdata[1],Latlngdata[2],"0000FF"])
		self.DistRadDict[str(ServiceRadLoc)] = "00CCFF"
		
		count  = 0
		self.DistGauge.SetValue(count)
		Task_Range = len(self.DemandLoc)
		self.DistGauge.SetRange(Task_Range)
		self.GaugeText.SetLabel('Status: Calculating Distances')
		
		#Geoencodes and adds Icons to the map
		for e in self.DemandLoc:
			time.sleep(0.2)
			Latlngdata = LinkBuilder.Geoencoding(e)
			if self.RadUnitSelect.GetValue() == "Minutes":
				GoogleData = LinkBuilder.CalculateDistance(e,ServiceRadLoc,"driving","Minutes")[0]
			elif self.RadUnitSelect.GetValue() == "Kilometres":
				GoogleData = LinkBuilder.CalculateDistance(e,ServiceRadLoc,"driving","Minutes","Kilometres")[1]
			elif self.RadUnitSelect.GetValue() == "Miles":
				GoogleData = LinkBuilder.CalculateDistance(e,ServiceRadLoc,"driving","Minutes","Miles")[1]
			
			self.DistRadDict[str(e)] = GoogleData
			
			count+=1
			self.DistGauge.SetValue(count)
			
			if Latlngdata[0] == "OK":
				if GoogleData < DrivingRadius:
					MappingData.append([str(e),Latlngdata[1],Latlngdata[2],"00CC33"])
					DemandWithin.append(str(e))
					self.DistRadAppend[str(e)] = ['True']
					if self.DemandSummary.GetValue() == True:
						#Saves the result
						outfile = open(fileloc,'ab')
						output=csv.writer(outfile)
						output.writerow([str(e),'True'])
						outfile.close()
				else:
					MappingData.append([str(e),Latlngdata[1],Latlngdata[2],"FF0000"])
					DemandOutside.append(str(e))
					self.DistRadAppend[str(e)] = ['False']
					if self.DemandSummary.GetValue() == True:
						#Saves the result
						outfile = open(fileloc,'ab')
						output=csv.writer(outfile)
						output.writerow([str(e),'False'])
						outfile.close()
		
		if self.AppendOutputSelect.GetValue() ==True:
			try:
				self.AppendData(ServiceRadLoc,self.DistRadAppend,1,Header)
			except:
				wx.MessageBox('The original data file could not be appended to, it may be open in another location or may no longer exist', 'Original File Not Found', wx.OK | wx.ICON_ERROR)

		if self.MapSave.GetValue() == True:
			LinkBuilder.StaticMapRadius(ServiceRadLoc,DemandWithin,DemandOutside,fileloc[:-4])
		if self.MapOutput.GetValue() == True: 
			self.GaugeText.SetLabel('Status: Creating Map')
			self.MappingData = MappingData
			
			#Creates the html file with markers and reloads the map
			mymap = pygmapscustom.maps(MappingData[0][1],MappingData[0][2],10,MappingData)
			mymap.draw('./Tempmap.html')
			TempMap = htmldir + "\Tempmap.html"
			self.browser.LoadURL(TempMap)
			self.DistanceRadSlider.SetValue(DrivingRadius)
			self.DistanceRadSlider.Enable()
		
		self.GaugeText.SetLabel('Status: Ready')

		
	def ChangeBySlider(self,event):
		
		if not self.MappingData == []:
			for e in self.MappingData:
				if self.DistRadDict[e[0]] == '00CCFF':
					e = e
				else:
					if self.DistRadDict[e[0]] < self.DistanceRadSlider.GetValue():
						e[3] = "339900"
					else:
						e[3] = "FF0000"
			
			mymap = pygmapscustom.maps(self.MappingData[0][1],self.MappingData[0][2],10,self.MappingData)
			mymap.draw('./Tempmap.html')
			TempMap = htmldir + "\Tempmap.html"
			self.browser.LoadURL(TempMap)			
			
	#Allows the user to select a column to take the data from
	def ChoosePostColumn(self):
		#Clears the box
		self.PostcodeColumnSelector.Clear()
		
		#Sets the drop down box to the headers of the selected file
		for e in self.list:
			self.PostcodeColumnSelector.Append(e)	
		
		self.PostcodeColumnSelector.SetValue(self.list[0])	
	
	#Changes the column to the one selected from the dropdown box
	def GetPostSel(self,event):
		if not self.list == [] :
			PostCol = "%s" %(self.list[event.GetSelection()])
			PostCol =  self.list.index(PostCol)
			self.ChoicePostCol = PostCol


	#Reads in the file with the currently selected postcode header
	def Rescan(self,event):
		self.DemandLoc = {}
		dummylist = []
		dummylist2 = []
		self.GaugeText.SetLabel('Status: Reading Files')
		#Reads CSV
		if '.csv' in self.DemandPath :
			infile = open(self.DemandPath,'rb')
			input=csv.reader(infile)
			data=[[i for i in row] for row in input]		
			infile.close()
			
			for row in data[1:]:
				if not row[self.ChoicePostCol] == "":
					dummylist.append(row[self.ChoicePostCol])
			
		#Reads XLSX files
		elif '.xlsx' in self.DemandPath:
			Demandwb = openpyxl.load_workbook(self.DemandPath)
			DemandSheet = Demandwb.get_active_sheet()
			Demandrows = DemandSheet.get_highest_row()

			for e in range(1,Demandrows):
					a = DemandSheet.cell(row = e,column = self.ChoicePostCol)
					if not a.value == None:
						dummylist.append(a.value)
		
		#Reads XLS files
		elif '.xls' in self.DemandPath[-4:]:
			DemandRows = []
			Demandwb = xlrd.open_workbook(self.DemandPath)
			DemandSheet = Demandwb.sheet_by_index(0)

			for rownum in range(1,DemandSheet.nrows):
				DemandRows.append(DemandSheet.row_values(rownum))
				
			for e in DemandRows:
				if not e[self.ChoicePostCol] == "":
					dummylist.append(str(e[self.ChoicePostCol]))
		
		for e in dummylist:
			dummylist2.append(e.encode('ascii', 'ignore').strip())
		dummylist = dummylist2
		
		for e in dummylist:
				if len(e)<12:
					if e in self.DemandLoc:
						self.DemandLoc[e] +=1
					else:
						self.DemandLoc[e] = 1
						
		self.ChoosePostButt.SetBackgroundColour('Green')
		self.GaugeText.SetLabel('Status: Ready')
	#This function allows files to be selected		
	def OpenFile(self):
	
		#This is defining a file dialogue
		if self.fileselect == 1 :
			Datatype = "Open the service location file"
		else:
			Datatype = "Open the Demand Location file"
			
		self.GaugeText.SetLabel('Status: Reading Files')
		
		input_stream = ""	
		openFileDialog = wx.FileDialog(self,Datatype, os.getcwd(),"" ,"All Files(*.*)|*.*|csv Files (*.csv)|*.csv|xls files (*.xls*)|*.xls*", wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)
		
		#Closes the dialog if the user cancels
		if openFileDialog.ShowModal() == wx.ID_CANCEL:
			return
		
		#Returns the file path
		input_stream = openFileDialog.GetPath()
		
		#Selects Service files
		if self.fileselect == 1 :
			#Ensures Empty Variables
			self.ServiceLoc = {}
			self.ServicePath = input_stream
			dummylist = []
			dummylist2 = []

			#Reads CSV data
			if '.csv' in self.ServicePath:
				infile = open(self.ServicePath,'rb')
				input=csv.reader(infile)
				
				for row in input:
					for i in row:
						if not i == "":
							dummylist.append(i)
						
				infile.close()
			
			#Reads XLSX data
			elif '.xlsx' in self.ServicePath:
				Servicewb = openpyxl.load_workbook(self.ServicePath)
				ServiceSheet = Servicewb.get_active_sheet()
				Servicerows = ServiceSheet.get_highest_row()
				Servicecols = ServiceSheet.get_highest_column()
				for e in range(Servicerows):
					for i in range(Servicecols):
						cell = ServiceSheet.cell(row = e, column = i)

						if not cell.value == None:
							dummylist.append(cell.value)
			
			#Reads XLS data
			elif '.xls' in self.ServicePath[-4:]:
				ServiceRows = []
				Servicewb = xlrd.open_workbook(self.ServicePath)
				ServiceSheet = Servicewb.sheet_by_index(0)
				
				for rownum in range(ServiceSheet.nrows):
					ServiceRows.append(ServiceSheet.row_values(rownum))
					
				for e in ServiceRows:
					for i in e:
						if not i == "":
							dummylist.append(str(i))
				
			#only allows supported file formats
			else:
				wx.MessageBox('Please Choose a Supported Data Format (.csv,.xlsx)', 'Invalid File Type', wx.OK | wx.ICON_ERROR)
				return
			
			#Removes Whitespace
			for e in dummylist:
				dummylist2.append(e.encode('ascii', 'ignore').strip())
			dummylist = dummylist2
			
			#Caches Service locations
			ServiceNo = 0
			for e in dummylist:
				ServiceNo += 1
				self.ServiceLoc[e] = ServiceNo
			
			self.ServiceOpen.SetBackgroundColour('Green')

		#Selects Demand files
		else:
			
			#Ensures Empty Variables
			self.DemandLoc = {}
			self.DemandPath = input_stream
			dummylist = []
			dummylist2 = []
			self.list = []
			
			#Reads CSV
			if '.csv' in self.DemandPath :
				infile = open(self.DemandPath,'rb')
				input=csv.reader(infile)
				data=[[i for i in row] for row in input]		
				infile.close()
				
				self.list = data[0]
			
			#Reads XLSX files
			elif '.xlsx' in self.DemandPath:
				Demandwb = openpyxl.load_workbook(self.DemandPath)
				DemandSheet = Demandwb.get_active_sheet()
				Demandrows = DemandSheet.get_highest_row()
				self.list = []
				for e in range(DemandSheet.get_highest_column()):
					a = DemandSheet.cell(row = 0, column = e)
					if not a.value == None:
						self.list.append("%s" %(a.value))
					else:
						self.list.append("")

			#Reads XLS files
			elif '.xls' in self.DemandPath[-4:]:
				DemandRows = []
				Demandwb = xlrd.open_workbook(self.DemandPath)
				DemandSheet = Demandwb.sheet_by_index(0)
				for e in DemandSheet.row_values(0):
					self.list.append(e)
				
			#only allows supported file formats
			else:
				wx.MessageBox('Please choose a supported data format (.csv,.xlsx)', 'Invalid file type', wx.OK | wx.ICON_ERROR)
				return
			
			self.ChoosePostColumn()
			self.DemandOpen.SetBackgroundColour('Green')
		self.GaugeText.SetLabel('Status: Ready')
	
	#Calculates the distance matrix	
	def Calculate(self,DataS,DataD):
		#Clears the variables
		RowData=[]
		NoRoutes = []
		BadPost = []
		MappingData = []
		
		for e in range(1,6):

			postcodestring = 'DataS[self.ServiceEntryLoc%s.GetValue()] = self.ServiceEntryName%s.GetValue()' %(e,e) 
			exec(postcodestring)
		
		if '' in DataS:
			del DataS['']

		#Ensures there is data to send
		if DataD == {} or DataS =={}:
			wx.MessageBox('Please select demand and service files', 'Missing files!', wx.OK | wx.ICON_ERROR)
			return
		#Sets up the progress gauge
		count = 0
		Task_Range = 0
		if self.DemandSummary.GetValue() == True or self.AppendOutputSelect.GetValue() == True:
			Task_Range += len(DataS)*len(DataD)
		
		if self.MapOutput.GetValue() == True:
			Task_Range +=len(DataD) + len(DataS)
		
		self.DistGauge.SetRange(Task_Range)
		self.DistGauge.SetValue(count)
		
		#Allows the user to enter a save location, or timestamps by default
		if self.UserSaveLoc.GetValue() == False:
			filebase = ".\Testdata\%s" %(time.time())
			filetype = ".csv"
			fileloc = filebase[:-3] + filetype
			
		else:
			fileloc = ""	
			SaveFileDialog = wx.FileDialog(self,"Choose save location", os.getcwd(),"" ,"csv Files (*.csv)|*.csv", wx.FD_SAVE)
		
			#Closes the dialog if the user cancels
			if SaveFileDialog.ShowModal() == wx.ID_CANCEL:
				return
		
			#Returns the file path
			fileloc = SaveFileDialog.GetPath()
		if self.DemandSummary.GetValue() == True or self.AppendOutputSelect.GetValue() == True:
			
			if self.DemandSummary.GetValue() == True:
				#Adds the unit of time to the data file
				RowData.append("%s %s" %(self.TimeUnitSelect.GetValue(),self.DistUnitSelect.GetValue()))
				RowData.append("Demand")
				
				#adds the demand locations as a heading
				for e in DataS:
					RowData.append(e + ' Time')
					RowData.append(e + ' Distance')

			if len(DataS)*len(DataD) < 2400:
				slowparam = 0.2
			else:
				slowparam = 35
			
			if self.DemandSummary.GetValue() == True:
			
			
				#Writes the header
				outfile = open(fileloc,'wb')
				output=csv.writer(outfile)		
			
				output.writerow(RowData)
					
				outfile.close()
			self.GaugeText.SetLabel('Status: Calculating Distances')     
			#Sends the requests, and returns the data
			for D in DataD:
				RowData = []
				RowData.append(D)
				RowData.append(DataD[D])
				
				for S in DataS:
					count+=1
					self.DistGauge.SetValue(count)
					time.sleep(slowparam)

					GoogleData = LinkBuilder.CalculateDistance(D,S,self.TravelMethod.GetValue().lower(),self.TimeUnitSelect.GetValue(),self.DistUnitSelect.GetValue())
					
					#Chooses the correct response to error messages
					if GoogleData == "OVER_QUERY_LIMIT":
						break
					
					if GoogleData[0] == "NOT_FOUND":
						BadPost.append(D)
					
					if GoogleData[0] == "ZERO_RESULTS":
						NoRoutes.append(D)
						
					RowData.append(GoogleData[0])
					RowData.append(GoogleData[1])
				
				self.FileAppendDict[D] = RowData[2:]
				
				#Saves the data upon each loop, to ensure no loss of data
				if self.DemandSummary.GetValue() == True:
					outfile = open(fileloc,'ab')
					output=csv.writer(outfile)
					output.writerow(RowData)
					outfile.close()	
				
				if GoogleData == "OVER_QUERY_LIMIT":
						break			
			
			#Informs the user that they have gone over the limit
			if GoogleData == "OVER_QUERY_LIMIT":
				wx.MessageBox('You have exceeded the request limit \nTry again later, although you may have to wait up to 24 hours to receive a new quota \nYour partial output has been saved', 'Too many requests', wx.OK | wx.ICON_ERROR)
			
			#Tells the user which postcodes didnt work
			if not NoRoutes == [] or not BadPost == []:
				BadPostString = ""
				NoRoutesString = ""
				for e in BadPost:
					BadPostString += "%s " %(e)
					
				for e in NoRoutes:
					NoRoutesString += "%s " %(e)

				BadDataString = "\tThere were several requests amongst the data: \n\tThese post codes could not be geoencoded \n\t%s \n\tThese postcodes could not be connected to service locations \n\t %s" %(BadPostString,NoRoutesString)
				wx.MessageBox(BadDataString, 'Unable to find data', wx.OK | wx.ICON_WARNING)
			
		if self.AppendOutputSelect.GetValue() == True:
			
			try:
				self.AppendData(DataS,self.FileAppendDict,0)		
			except:
				wx.MessageBox('The original data file could not be appended to, it may be open in another location or may no longer exist', 'Original File Not Found', wx.OK | wx.ICON_ERROR)
			
		MaxDemand = DataD.itervalues().next()
		for e in DataD:
			if DataD[e] > MaxDemand:
				MaxDemand = DataD[e]
		
		if self.MapSave.GetValue() == True:
				if self.MapType.GetValue() == "Standard":
					LinkBuilder.StaticMaps(DataS,DataD,fileloc[:-4])
				elif self.MapType.GetValue() == "Heat Map":
				
					for e in range(11):
						self.HeatDict[e] = []
						for i in DataD:
							if int(10*DataD[str(i)]/MaxDemand) == e:
								self.HeatDict[e].append(i)
					
					LinkBuilder.StaticMapHeat(DataS,self.HeatDict,self.hexdict,fileloc[:-4])
		if self.MapOutput.GetValue() == True:
				
			self.GaugeText.SetLabel('Status: Geoencoding')
			#Geoencodes the data
			for e in DataS:
				count +=1
				self.DistGauge.SetValue(count)
				
				time.sleep(0.2)
				Latlngdata = LinkBuilder.Geoencoding(e)
				
				if Latlngdata[0] == "OK": 
					MappingData.append([str(e).upper() + ' %s' %(str(DataS[e])),Latlngdata[1],Latlngdata[2],"0000FF"])
					
			for e in DataD:
				count +=1
				self.DistGauge.SetValue(count)
				time.sleep(0.2)
				
				Latlngdata = LinkBuilder.Geoencoding(e)
				
				if Latlngdata[0] == "OK": 
					if self.MapType.GetValue() == "Heat Map":
						MappingData.append([str(e).upper() + ' Demand : %s' %(self.DemandLoc[str(e)]),Latlngdata[1],Latlngdata[2],self.hexdict[int(10*DataD[str(e)]/MaxDemand)]])
					else:
						MappingData.append([str(e).upper() + ' Demand : %s' %(self.DemandLoc[str(e)]),Latlngdata[1],Latlngdata[2],'FFFF00'])
			
			#Creates the html file with markers and reloads the map
			mymap = pygmapscustom.maps(MappingData[0][1],MappingData[0][2],10,MappingData)
			mymap.draw('./Tempmap.html')
			TempMap = htmldir + "\Tempmap.html"
			self.browser.LoadURL(TempMap)
			
		#Shows the user that the Tool is ready
		self.GaugeText.SetLabel('Status: Ready')
	
	#Creates a new file with the distance matrix data appended
	def AppendData(self,DataS,AppendData,select,HeaderRad = False):
		self.GaugeText.SetLabel('Status: Altering Original File')
		#Creates a copy of the original demand file
		PathwoEXT = os.path.splitext(self.DemandPath)
		CopyFilePath = PathwoEXT[0] + '_Append' + '.' + self.DemandPath.split('.')[1]
		CopyFilePath = str(CopyFilePath)
		shutil.copyfile(self.DemandPath,CopyFilePath)
		
		#Creates the header
		Header = []
		if select == 0:
			for e in DataS:
				Header.append(e + ' Time')
				Header.append(e + ' Distance')
		elif select == 1:
			Header.append(HeaderRad)
			
		#Outputs the modified file
		if '.csv' in CopyFilePath:
			
			infile = open(CopyFilePath,'rb')
			input=csv.reader(infile)
			data=[[i for i in row] for row in input]		
			infile.close()
			
			data[0] += Header
			for e in data[1:]:
				try:
					e += AppendData[e[self.ChoicePostCol]]
				except:
					e += ['ERROR','ERROR']
			
			outfile = open(CopyFilePath,'wb')
			output = csv.writer(outfile)
			for e in data:
				output.writerow(e)
			outfile.close()			
			
		elif '.xlsx' in CopyFilePath:
			AppendWB = openpyxl.load_workbook(CopyFilePath)
			AppendWS = AppendWB.get_active_sheet()
			data = [[i.value for i in row] for row in AppendWS.rows]
			print data
			print AppendData
			print 'woof'
			data[0] +=Header
			for e in data[1:]:
				try:
					e += AppendData[e[self.ChoicePostCol]]
				except:
					e += ['ERROR','ERROR']
			for e in range(len(data)):
				for i in range(len(data[e])):
					AppendWS.cell(row = e,column = i).value = data[e][i]
					
			AppendWB.save(CopyFilePath)	
			
		elif '.xls' in CopyFilePath[-4:]:
			AppendWB = xlrd.open_workbook(CopyFilePath)
			AppendWS = AppendWB.sheet_by_index(0)
			data = []
			
			for rownum in range(0,AppendWS.nrows):
				data.append(AppendWS.row_values(rownum))
				
			data[0] += Header
			for e in data[1:]:
				try:
					e += AppendData[e[self.ChoicePostCol]]
				except:
					e += ['ERROR','ERROR']
			
			AppendWB = xlwt.Workbook()
			AppendWS = AppendWB.add_sheet('Sheet 1')
			
			
			for e in range(len(data)):
				for i in range(len(data[e])):
					AppendWS.write(e,i,data[e][i])
					
			AppendWB.save(CopyFilePath)
#This class is the Gui		
class MyApp(wx.App):
	def OnInit(self):
		frame = MainFrame(None,-1,'PatMap')
			
		frame.Show(True)
		frame.Maximize()
		
		welcomewindow = Welcome.WelcomeScreen(None)
		welcomewindow.Show()
		
		loc = wx.IconLocation(htmldir + '\GIStool.ico', 0)
		frame.SetIcon(wx.IconFromLocation(loc))
		frame.Centre()
		return True

	
#Starts the GUI
app = MyApp(0)
app.MainLoop()
